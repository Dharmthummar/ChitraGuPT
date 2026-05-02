from __future__ import annotations

import argparse
import base64
import json
import mimetypes
import os
import re
import secrets
import socket
import subprocess
import sys
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
import webbrowser
from collections import deque
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from flask import Flask, jsonify, render_template, request
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


SOURCE_DIR = Path(__file__).resolve().parent
RESOURCE_DIR = Path(getattr(sys, "_MEIPASS", SOURCE_DIR))
APP_DIR = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else SOURCE_DIR
DATA_DIR = APP_DIR / "data"
CONFIG_FILE = DATA_DIR / "config.json"
HISTORY_FILE = DATA_DIR / "history.jsonl"

DEFAULT_MODEL = "gemini-3.1-flash-lite-preview"
GEMINI_FALLBACK_MODELS = (
    DEFAULT_MODEL,
    "gemini-3-flash-preview",
    "gemini-2.5-flash-lite",
    "gemini-2.5-flash",
)
GEMINI_API_VERSION = "v1beta"
GEMINI_RETRY_DELAYS_SECONDS = (2, 5, 10)
MAX_INLINE_BYTES = 18 * 1024 * 1024
MAX_REQUEST_BYTES = 24 * 1024 * 1024
MAX_HEADER_SCAN_ROWS = 40
SUPPORTED_EXCEL = {".xlsx", ".xlsm"}
SUPPORTED_UPLOAD_MIME = {
    "application/pdf",
    "image/jpeg",
    "image/png",
    "image/webp",
    "image/heic",
    "image/heif",
}

DATA_DIR.mkdir(parents=True, exist_ok=True)
CONFIG_LOCK = threading.Lock()
HISTORY_LOCK = threading.Lock()
INSPECT_CACHE_LOCK = threading.Lock()
INSPECT_CACHE: dict[tuple[str, int, int, str], dict[str, Any]] = {}
MAX_INSPECT_CACHE_ITEMS = 24
WORKBOOK_LOCKS_LOCK = threading.Lock()
WORKBOOK_LOCKS: dict[str, threading.Lock] = {}

app = Flask(
    __name__,
    static_folder=str(RESOURCE_DIR / "static"),
    template_folder=str(RESOURCE_DIR / "templates"),
)
app.config["MAX_CONTENT_LENGTH"] = MAX_REQUEST_BYTES
app.config["JSON_SORT_KEYS"] = False
app.config.setdefault("PUBLIC_BASE_URL", "")
app.config.setdefault("PUBLIC_SHARE_TOKEN", "")
app.config.setdefault("PUBLIC_SHARE_ERROR", "")

PUBLIC_SHARE_HEADER = "X-Chitra-Share"
FORWARDED_HEADER_NAMES = ("Cf-Connecting-Ip", "X-Forwarded-For", "X-Real-Ip")


class GeminiApiError(RuntimeError):
    def __init__(self, status_code: int, message: str) -> None:
        super().__init__(message)
        self.status_code = status_code


def set_public_share(base_url: str, token: str) -> None:
    app.config["PUBLIC_BASE_URL"] = base_url.rstrip("/")
    app.config["PUBLIC_SHARE_TOKEN"] = token
    app.config["PUBLIC_SHARE_ERROR"] = ""


def set_public_share_error(message: str) -> None:
    app.config["PUBLIC_SHARE_ERROR"] = message


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def load_config() -> dict[str, Any]:
    default = {
        "provider": "gemini",
        "gemini_api_key": os.environ.get("GEMINI_API_KEY", ""),
        "gemini_model": DEFAULT_MODEL,
        "default_excel_path": "",
        "default_sheet": "",
        "recent_sheets": [],
    }
    if not CONFIG_FILE.exists():
        return default

    try:
        data = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return default

    merged = {**default, **data}
    if not merged.get("gemini_api_key"):
        merged["gemini_api_key"] = os.environ.get("GEMINI_API_KEY", "")
    merged["gemini_model"] = clean_model_name(merged.get("gemini_model", DEFAULT_MODEL))
    return merged


def save_config(config: dict[str, Any]) -> None:
    payload = json.dumps(config, indent=2)
    temp_file = CONFIG_FILE.with_suffix(".tmp")
    with CONFIG_LOCK:
        temp_file.write_text(payload, encoding="utf-8")
        temp_file.replace(CONFIG_FILE)


def display_model_name(model: str) -> str:
    return display_from_resource_name(model)


def public_config(config: dict[str, Any]) -> dict[str, Any]:
    return {
        "provider": config.get("provider", "gemini"),
        "hasApiKey": bool(config.get("gemini_api_key")),
        "geminiModel": display_model_name(config.get("gemini_model", DEFAULT_MODEL)),
        "defaultExcelPath": config.get("default_excel_path", ""),
        "defaultSheet": config.get("default_sheet", ""),
        "recentSheets": config.get("recent_sheets", [])[:8],
    }


def normalize_path(path_text: str) -> Path:
    expanded = os.path.expandvars(path_text.strip().strip('"'))
    return Path(expanded).expanduser()


def require_excel_path(path_text: str) -> Path:
    if not path_text or not path_text.strip():
        raise ValueError("Excel workbook path is required.")

    path = normalize_path(path_text)
    if not path.exists():
        raise FileNotFoundError(f"Workbook was not found: {path}")
    if path.suffix.lower() not in SUPPORTED_EXCEL:
        raise ValueError("Use an .xlsx or .xlsm workbook. Old .xls files cannot be updated safely.")
    return path


def workbook_kwargs(path: Path) -> dict[str, Any]:
    return {"keep_vba": path.suffix.lower() == ".xlsm"}


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def workbook_signature(path: Path) -> tuple[str, int, int]:
    stat = path.stat()
    return (str(path.resolve()), stat.st_mtime_ns, stat.st_size)


def workbook_lock(path: Path) -> threading.Lock:
    key = str(path.resolve())
    with WORKBOOK_LOCKS_LOCK:
        lock = WORKBOOK_LOCKS.get(key)
        if lock is None:
            lock = threading.Lock()
            WORKBOOK_LOCKS[key] = lock
        return lock


def score_header_cells(values: list[str]) -> int:
    non_empty = [value for value in values if value]
    if len(non_empty) < 2:
        return 0

    unique_count = len({value.lower() for value in non_empty})
    alpha_count = sum(bool(re.search(r"[A-Za-z]", value)) for value in non_empty)
    short_text_count = sum(len(value) <= 60 for value in non_empty)
    long_text_count = sum(len(value) > 80 for value in non_empty)
    mostly_numeric_count = sum(bool(re.fullmatch(r"[$₹€£,.\d\s:/\\-]+", value)) for value in non_empty)

    score = len(non_empty) * 4 + unique_count * 2 + alpha_count * 3 + short_text_count
    score -= mostly_numeric_count * 4
    score -= long_text_count * 3

    if alpha_count < max(2, len(non_empty) // 2):
        score -= 8
    return score


def detect_header_row(ws) -> tuple[int, list[str]]:
    best_row_number = 1
    best_headers: list[str] = []
    best_score = -1
    max_scan_row = min(ws.max_row or 1, MAX_HEADER_SCAN_ROWS)

    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_row, values_only=True),
        start=1,
    ):
        headers = [cell_to_text(value) for value in row]
        score = score_header_cells(headers)
        if score > best_score:
            best_score = score
            best_row_number = row_number
            best_headers = headers

    return best_row_number, best_headers


def last_data_row_for_headers(ws, header_row_number: int, headers: list[str]) -> int:
    header_indexes = [index + 1 for index, header in enumerate(headers) if header]
    if not header_indexes:
        return header_row_number

    for row_number in range(ws.max_row or header_row_number, header_row_number, -1):
        if any(cell_to_text(ws.cell(row=row_number, column=col_index).value) for col_index in header_indexes):
            return row_number
    return header_row_number


def inspect_workbook(path_text: str, sheet_name: str | None = None) -> dict[str, Any]:
    path = require_excel_path(path_text)
    requested_sheet = sheet_name or ""
    cache_key = (*workbook_signature(path), requested_sheet)
    with INSPECT_CACHE_LOCK:
        cached = INSPECT_CACHE.get(cache_key)
        if cached is not None:
            return json.loads(json.dumps(cached))

    wb = load_workbook(path, read_only=True, data_only=True, **workbook_kwargs(path))
    try:
        sheet_names = wb.sheetnames
        selected = sheet_name if sheet_name in sheet_names else sheet_names[0]
        ws = wb[selected]

        header_row_number, headers = detect_header_row(ws)
        valid_headers = [header for header in headers if header]
        last_data_row = last_data_row_for_headers(ws, header_row_number, headers)

        samples: list[dict[str, str]] = []
        start_row = max(header_row_number + 1, last_data_row - 30)
        for row in ws.iter_rows(min_row=start_row, max_row=last_data_row, values_only=True):
            values = [cell_to_text(value) for value in row]
            if not any(values):
                continue
            sample = {}
            for index, header in enumerate(headers):
                if header and index < len(values):
                    sample[header] = values[index]
            if sample:
                samples.append(sample)

        result = {
            "path": str(path),
            "fileName": path.name,
            "sheetNames": sheet_names,
            "sheet": selected,
            "headerRow": header_row_number,
            "lastDataRow": last_data_row,
            "headers": headers,
            "validHeaders": valid_headers,
            "rowCount": max(last_data_row - header_row_number, 0),
            "sampleRows": samples[-3:],
            "checkedAt": now_iso(),
        }
        with INSPECT_CACHE_LOCK:
            INSPECT_CACHE[cache_key] = result
            if len(INSPECT_CACHE) > MAX_INSPECT_CACHE_ITEMS:
                oldest_key = next(iter(INSPECT_CACHE))
                INSPECT_CACHE.pop(oldest_key, None)
        return json.loads(json.dumps(result))
    finally:
        wb.close()


def remember_sheet(path_text: str, sheet: str) -> None:
    config = load_config()
    path = str(normalize_path(path_text))
    recent = config.get("recent_sheets", [])
    recent = [
        item for item in recent
        if not (item.get("path") == path and item.get("sheet") == sheet)
    ]
    recent.insert(0, {"path": path, "sheet": sheet, "lastUsed": now_iso()})
    config["recent_sheets"] = recent[:8]
    config["default_excel_path"] = path
    config["default_sheet"] = sheet
    save_config(config)


def clean_model_name(name: str) -> str:
    """Return a safe Gemini REST model resource name."""
    candidate = str(name or DEFAULT_MODEL).strip()
    if candidate.startswith("models/"):
        candidate = candidate.removeprefix("models/")
    if not re.fullmatch(r"[A-Za-z0-9._-]+", candidate):
        candidate = DEFAULT_MODEL
    return f"models/{candidate}"


def display_from_resource_name(model: str) -> str:
    return clean_model_name(model).removeprefix("models/")


def gemini_model_candidates(model: str) -> list[str]:
    preferred = display_from_resource_name(model)
    candidates = [preferred, *GEMINI_FALLBACK_MODELS]
    unique: list[str] = []
    for candidate in candidates:
        cleaned = display_from_resource_name(candidate)
        if cleaned not in unique:
            unique.append(cleaned)
    return [f"models/{candidate}" for candidate in unique]


def is_model_unavailable_error(error: GeminiApiError) -> bool:
    message = str(error).lower()
    if error.status_code not in {400, 403, 404}:
        return False
    return "model" in message and any(
        phrase in message
        for phrase in (
            "not found",
            "not supported",
            "not available",
            "deprecated",
            "permission",
            "does not exist",
        )
    )


def build_prompt(headers: list[str], sample_rows: list[dict[str, str]]) -> str:
    valid_headers = [header for header in headers if header]
    return f"""
You extract data from an uploaded business document into an existing Excel row.

Return exactly one JSON object. No markdown, no commentary.

Excel headers, in exact order:
{json.dumps(valid_headers, ensure_ascii=False)}

Recent rows from the same sheet:
{json.dumps(sample_rows, ensure_ascii=False)}

Rules:
- Use exactly the same keys as the Excel headers.
- Use an empty string for a field that is not visible in the document.
- Match the existing row style for dates, invoice numbers, taxes, totals, and currency.
- If the document has multiple line items, create one summary row using the invoice total.
- Do not invent values.
- Do not include keys that are not Excel headers.
""".strip()


def parse_json_object(text: str) -> dict[str, Any]:
    cleaned = text.strip()
    cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*```$", "", cleaned)

    try:
        value = json.loads(cleaned)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", cleaned, flags=re.DOTALL)
        if not match:
            raise ValueError("The AI response did not contain a JSON object.")
        value = json.loads(match.group(0))

    if isinstance(value, list):
        if not value or not isinstance(value[0], dict):
            raise ValueError("The AI response was JSON, but not a row object.")
        value = value[0]

    if not isinstance(value, dict):
        raise ValueError("The AI response was JSON, but not a row object.")

    return value


def normalize_row(row_data: dict[str, Any], headers: list[str]) -> dict[str, Any]:
    exact_keys = {key: key for key in row_data}
    lower_keys = {key.lower().strip(): key for key in row_data}
    normalized: dict[str, Any] = {}

    for header in headers:
        if not header:
            continue
        source_key = exact_keys.get(header) or lower_keys.get(header.lower().strip())
        value = row_data.get(source_key, "") if source_key else ""
        if value is None:
            value = ""
        elif isinstance(value, (list, dict)):
            value = json.dumps(value, ensure_ascii=False)
        normalized[header] = value

    return normalized


def format_gemini_http_error(status_code: int, body: str) -> str:
    try:
        error_json = json.loads(body)
        message = error_json.get("error", {}).get("message", body)
    except json.JSONDecodeError:
        message = body

    if status_code == 429:
        model_match = re.search(r"model:\s*([^\n,*]+)", message)
        model_label = model_match.group(1).strip() if model_match else DEFAULT_MODEL

        if "limit: 0" in message:
            return (
                f"Gemini API quota is not enabled for {model_label} on this Google project. "
                "Check the active limits at https://ai.dev/rate-limit, or set up billing in "
                "Google AI Studio if you need quota for this model."
            )

        retry_match = re.search(r"retry in ([\d.]+)s", message, flags=re.IGNORECASE)
        retry_note = f" Wait about {retry_match.group(1)} seconds and retry." if retry_match else ""
        return (
            f"Gemini API rate limit reached for {model_label}."
            f"{retry_note} You can monitor active limits at https://ai.dev/rate-limit."
        )

    if status_code == 503:
        return (
            "Gemini is temporarily experiencing high demand for this model. "
            "The app retried automatically, but Google still did not have capacity. "
            "Wait a minute, then run extraction again."
        )

    return f"Gemini API error {status_code}: {message}"


def call_gemini(
    api_key: str,
    model: str,
    file_bytes: bytes,
    mime_type: str,
    headers: list[str],
    sample_rows: list[dict[str, str]],
) -> tuple[dict[str, Any], str]:
    if not api_key:
        raise RuntimeError("Add a Gemini API key in Settings first.")
    if len(file_bytes) > MAX_INLINE_BYTES:
        raise ValueError("File is too large for inline AI reading. Use a smaller file under 18 MB.")

    prompt = build_prompt(headers, sample_rows)
    encoded_file = base64.b64encode(file_bytes).decode("ascii")
    payload = {
        "contents": [
            {
                "parts": [
                    {"inline_data": {"mime_type": mime_type, "data": encoded_file}},
                    {"text": prompt},
                ]
            }
        ],
        "generationConfig": {
            "temperature": 0,
            "maxOutputTokens": 2048,
            "responseMimeType": "application/json",
        },
    }

    def post_payload(model_name: str, request_payload: dict[str, Any]) -> dict[str, Any]:
        url = f"https://generativelanguage.googleapis.com/{GEMINI_API_VERSION}/{model_name}:generateContent"
        request_body = json.dumps(request_payload, separators=(",", ":")).encode("utf-8")
        req = urllib.request.Request(
            url,
            data=request_body,
            headers={
                "Content-Type": "application/json",
                "x-goog-api-key": api_key,
            },
            method="POST",
        )

        for attempt in range(len(GEMINI_RETRY_DELAYS_SECONDS) + 1):
            try:
                with urllib.request.urlopen(req, timeout=90) as response:
                    return json.loads(response.read().decode("utf-8"))
            except urllib.error.HTTPError as error:
                body = error.read().decode("utf-8", errors="replace")
                api_error = GeminiApiError(error.code, format_gemini_http_error(error.code, body))
                if error.code == 503 and attempt < len(GEMINI_RETRY_DELAYS_SECONDS):
                    time.sleep(GEMINI_RETRY_DELAYS_SECONDS[attempt])
                    continue
                raise api_error from error
            except urllib.error.URLError as error:
                reason = getattr(error, "reason", error)
                raise RuntimeError(f"Could not reach Gemini API: {reason}") from error

        raise RuntimeError("Gemini request failed after automatic retries.")

    response_json: dict[str, Any] | None = None
    used_model = ""
    last_model_error: GeminiApiError | None = None

    for candidate_model in gemini_model_candidates(model):
        try:
            response_json = post_payload(candidate_model, payload)
            used_model = candidate_model
            break
        except GeminiApiError as error:
            if "responseMimeType" in str(error) or "response_mime_type" in str(error):
                fallback_payload = {**payload, "generationConfig": {"temperature": 0, "maxOutputTokens": 2048}}
                response_json = post_payload(candidate_model, fallback_payload)
                used_model = candidate_model
                break
            if is_model_unavailable_error(error):
                last_model_error = error
                continue
            raise

    if response_json is None:
        if last_model_error:
            raise last_model_error
        raise RuntimeError("Gemini request failed because no compatible model was available.")

    parts = (
        response_json.get("candidates", [{}])[0]
        .get("content", {})
        .get("parts", [])
    )
    text = "".join(part.get("text", "") for part in parts).strip()
    if not text:
        raise RuntimeError("Gemini did not return extracted text.")

    parsed = parse_json_object(text)
    return normalize_row(parsed, headers), display_from_resource_name(used_model)


def append_to_workbook(
    path_text: str,
    sheet: str,
    headers: list[str],
    row_data: dict[str, Any],
    header_row_number: int | None = None,
) -> int:
    path = require_excel_path(path_text)
    with workbook_lock(path):
        wb = load_workbook(path, **workbook_kwargs(path))
        try:
            if sheet not in wb.sheetnames:
                raise ValueError(f"Sheet was not found: {sheet}")

            ws = wb[sheet]
            if header_row_number is None:
                header_row_number, detected_headers = detect_header_row(ws)
                if headers == [header for header in detected_headers[:len(headers)]]:
                    headers = detected_headers

            header_row_number = max(1, int(header_row_number or 1))
            last_data_row = last_data_row_for_headers(ws, header_row_number, headers)
            target_row = last_data_row + 1
            source_row = max(last_data_row, header_row_number)
            highlight = PatternFill(fill_type="solid", fgColor="DFF7EA")
            text_color = Font(color="0F3D2E")

            for col_index, header in enumerate(headers, start=1):
                source_cell = ws.cell(row=source_row, column=col_index)
                target_cell = ws.cell(row=target_row, column=col_index)
                if target_row > 2:
                    target_cell._style = copy(source_cell._style)
                    target_cell.number_format = source_cell.number_format
                    target_cell.alignment = copy(source_cell.alignment)
                    target_cell.border = copy(source_cell.border)
                target_cell.value = row_data.get(header, "") if header else ""
                if header:
                    target_cell.fill = highlight
                    target_cell.font = copy(text_color)

            wb.save(path)
            return target_row
        finally:
            wb.close()


def add_history(entry: dict[str, Any]) -> None:
    with HISTORY_LOCK:
        with HISTORY_FILE.open("a", encoding="utf-8") as history:
            history.write(json.dumps(entry, ensure_ascii=False) + "\n")


def read_history(limit: int = 20) -> list[dict[str, Any]]:
    if not HISTORY_FILE.exists():
        return []
    try:
        with HISTORY_FILE.open("r", encoding="utf-8") as history:
            lines = deque(history, maxlen=limit)
    except OSError:
        return []
    rows: list[dict[str, Any]] = []
    for line in reversed(lines):
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return rows


def detect_mime(filename: str, browser_mime: str | None) -> str:
    if browser_mime and browser_mime in SUPPORTED_UPLOAD_MIME:
        return browser_mime
    guessed, _ = mimetypes.guess_type(filename)
    if guessed in SUPPORTED_UPLOAD_MIME:
        return guessed
    suffix = Path(filename).suffix.lower()
    fallback = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".webp": "image/webp",
        ".pdf": "application/pdf",
        ".heic": "image/heic",
        ".heif": "image/heif",
    }.get(suffix)
    if fallback:
        return fallback
    raise ValueError("Upload a PDF or image file.")


def get_lan_ip() -> str:
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        sock.connect(("8.8.8.8", 80))
        return sock.getsockname()[0]
    except OSError:
        try:
            return socket.gethostbyname(socket.gethostname())
        except OSError:
            return "127.0.0.1"
    finally:
        sock.close()


def with_query_params(url: str, params: list[tuple[str, str]]) -> str:
    parts = urllib.parse.urlsplit(url)
    query = urllib.parse.parse_qsl(parts.query, keep_blank_values=True)
    query.extend((key, value) for key, value in params if value)
    return urllib.parse.urlunsplit((
        parts.scheme,
        parts.netloc,
        parts.path,
        urllib.parse.urlencode(query),
        parts.fragment,
    ))


def share_payload() -> dict[str, str]:
    host = request.host
    port = host.rsplit(":", 1)[1] if ":" in host else "5055"
    lan_url = f"http://{get_lan_ip()}:{port}"
    local_phone_url = with_query_params(lan_url, [("phone", "1")])
    current_url = request.host_url.rstrip("/")

    public_url = str(app.config.get("PUBLIC_BASE_URL") or os.environ.get("CHITRAGUPT_PUBLIC_URL", "")).strip().rstrip("/")
    share_token = str(app.config.get("PUBLIC_SHARE_TOKEN") or os.environ.get("CHITRAGUPT_SHARE_TOKEN", "")).strip()
    phone_url = local_phone_url
    if public_url:
        token_params = [("phone", "1")]
        if share_token:
            token_params.append(("share", share_token))
        phone_url = with_query_params(public_url, token_params)

    message = f"Upload invoice/photo here: {phone_url}"
    return {
        "currentUrl": current_url,
        "lanUrl": lan_url,
        "localPhoneUrl": local_phone_url,
        "publicUrl": public_url,
        "publicError": str(app.config.get("PUBLIC_SHARE_ERROR") or ""),
        "phoneUrl": phone_url,
        "mode": "public" if public_url else "lan",
        "whatsappUrl": "https://wa.me/?text=" + urllib.parse.quote(message),
        "emailUrl": "mailto:?subject=Upload invoice&body=" + urllib.parse.quote(message),
    }


def open_file_on_host(path_text: str) -> None:
    path = require_excel_path(path_text)
    if sys.platform.startswith("win"):
        script = r"""
$Path = $args[0]
Start-Process -FilePath $Path
Start-Sleep -Milliseconds 900
$shell = New-Object -ComObject WScript.Shell
$leaf = [System.IO.Path]::GetFileName($Path)
if (-not $shell.AppActivate($leaf)) {
  [void]$shell.AppActivate('Excel')
}
"""
        subprocess.Popen(
            [
                "powershell",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-Command",
                script,
                str(path),
            ],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
    elif sys.platform == "darwin":
        subprocess.Popen(["open", str(path)])
    else:
        subprocess.Popen(["xdg-open", str(path)])


def browse_excel_on_host() -> str:
    if sys.platform.startswith("win"):
        script = r"""
Add-Type -AssemblyName System.Windows.Forms
[Console]::OutputEncoding = [System.Text.Encoding]::UTF8
$owner = New-Object System.Windows.Forms.Form
$owner.Text = 'Connect Excel workbook'
$owner.TopMost = $true
$owner.ShowInTaskbar = $false
$owner.StartPosition = 'CenterScreen'
$owner.Width = 1
$owner.Height = 1
$owner.Opacity = 0.01
$dialog = New-Object System.Windows.Forms.OpenFileDialog
$dialog.Title = 'Connect Excel workbook'
$dialog.Filter = 'Excel workbooks (*.xlsx;*.xlsm)|*.xlsx;*.xlsm'
$dialog.Multiselect = $false
$dialog.CheckFileExists = $true
try {
  [void]$owner.Show()
  [void]$owner.Activate()
  if ($dialog.ShowDialog($owner) -eq [System.Windows.Forms.DialogResult]::OK) {
    Write-Output $dialog.FileName
  }
} finally {
  $owner.Close()
  $owner.Dispose()
}
"""
        try:
            result = subprocess.run(
                [
                    "powershell",
                    "-NoProfile",
                    "-STA",
                    "-ExecutionPolicy",
                    "Bypass",
                    "-Command",
                    script,
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout.strip().splitlines()[-1]
        except Exception as error:
            print(f"DEBUG: PowerShell browse failed: {error}")
            # Fallback to tkinter below
            pass

    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.askopenfilename(
            title="Connect Excel workbook",
            filetypes=[("Excel workbooks", "*.xlsx *.xlsm")],
        )
        root.destroy()
        return path
    except Exception as error:
        print(f"DEBUG: Tkinter browse failed: {error}")
        raise RuntimeError("This system cannot open a native file picker.") from error


def api_error(error: Exception, status: int = 400):
    return jsonify({"ok": False, "error": str(error)}), status


def is_public_tunnel_request() -> bool:
    if not app.config.get("PUBLIC_SHARE_TOKEN"):
        return False
    return any(request.headers.get(name) for name in FORWARDED_HEADER_NAMES)


def has_public_share_token() -> bool:
    expected = str(app.config.get("PUBLIC_SHARE_TOKEN") or "")
    supplied = request.args.get("share", "") or request.headers.get(PUBLIC_SHARE_HEADER, "")
    return bool(expected and supplied and secrets.compare_digest(supplied, expected))


@app.before_request
def require_public_share_token():
    if not is_public_tunnel_request():
        return None
    if request.path.startswith("/static/") or request.path in {"/favicon.ico", "/healthz"}:
        return None
    if has_public_share_token():
        return None

    message = "This public link is missing or expired. Copy a fresh Host link from the desktop app."
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "error": message}), 403
    return message, 403


@app.get("/")
def index():
    return render_template("index.html")


@app.get("/healthz")
def healthz():
    return jsonify({"ok": True, "time": now_iso()})


@app.get("/api/state")
def api_state():
    config = load_config()
    return jsonify({
        "ok": True,
        "config": public_config(config),
        "history": read_history(20),
        "share": share_payload(),
    })


@app.post("/api/settings")
def api_settings():
    try:
        payload = request.get_json(force=True) or {}
        config = load_config()
        api_key = str(payload.get("apiKey", "")).strip()
        model = clean_model_name(str(payload.get("model", "") or DEFAULT_MODEL).strip())

        if api_key:
            config["gemini_api_key"] = api_key
        config["gemini_model"] = model
        save_config(config)
        return jsonify({"ok": True, "config": public_config(config)})
    except Exception as error:
        return api_error(error)


@app.post("/api/inspect-excel")
def api_inspect_excel():
    path = ""
    try:
        payload = request.get_json(force=True) or {}
        path = str(payload.get("path", ""))
        sheet = str(payload.get("sheet", "") or "")
        info = inspect_workbook(path, sheet or None)
        if not info["validHeaders"]:
            raise ValueError("Row 1 must contain column headers.")
        remember_sheet(info["path"], info["sheet"])
        return jsonify({"ok": True, "sheet": info, "config": public_config(load_config())})
    except Exception as error:
        print(f"DEBUG: Inspection failed for {path}: {error}")
        return api_error(error)


@app.post("/api/browse-excel")
def api_browse_excel():
    try:
        selected_path = browse_excel_on_host()
        if not selected_path:
            return jsonify({"ok": True, "selected": False})

        info = inspect_workbook(selected_path)
        if not info["validHeaders"]:
            raise ValueError("Row 1 must contain column headers.")
        remember_sheet(info["path"], info["sheet"])
        return jsonify({
            "ok": True,
            "selected": True,
            "path": info["path"],
            "sheet": info,
            "config": public_config(load_config()),
        })
    except Exception as error:
        return api_error(error)


@app.post("/api/extract")
def api_extract():
    started = time.time()
    try:
        config = load_config()
        path = request.form.get("excelPath", "")
        sheet = request.form.get("sheet", "")
        try:
            header_row_number = int(request.form.get("headerRow", "") or 0) or None
        except ValueError:
            header_row_number = None
        uploaded = request.files.get("document")
        if not uploaded or not uploaded.filename:
            raise ValueError("Choose a PDF or image first.")
        if not path or not sheet:
            raise ValueError("Inspect the Excel sheet before running AI extraction.")

        mime_type = detect_mime(uploaded.filename, uploaded.mimetype)
        file_bytes = uploaded.read()
        if not file_bytes:
            raise ValueError("Uploaded file is empty.")
        if len(file_bytes) > MAX_INLINE_BYTES:
            raise ValueError("File is too large for inline AI reading. Use a smaller file under 18 MB.")
        
        # Performance Optimization: Use client headers to skip an extra Excel load
        headers_raw = request.form.get("headers")
        sample_rows_raw = request.form.get("sampleRows")
        
        if headers_raw:
            headers = json.loads(headers_raw)
            sample_rows = json.loads(sample_rows_raw or "[]")
        else:
            # Fallback for old clients or manual API calls
            sheet_info = inspect_workbook(path, sheet)
            headers = sheet_info["headers"]
            sample_rows = sheet_info["sampleRows"]
            header_row_number = int(sheet_info.get("headerRow") or 1)

        configured_model = config.get("gemini_model", DEFAULT_MODEL)
        row_data, model_used = call_gemini(
            api_key=config.get("gemini_api_key", ""),
            model=configured_model,
            file_bytes=file_bytes,
            mime_type=mime_type,
            headers=headers,
            sample_rows=sample_rows,
        )
        if clean_model_name(model_used) != clean_model_name(configured_model):
            config["gemini_model"] = clean_model_name(model_used)
            save_config(config)

        row_number = append_to_workbook(path, sheet, headers, row_data, header_row_number)
        remember_sheet(path, sheet)

        non_empty = {key: value for key, value in row_data.items() if str(value).strip()}
        entry = {
            "time": now_iso(),
            "fileName": uploaded.filename,
            "excelPath": str(normalize_path(path)),
            "sheet": sheet,
            "rowNumber": row_number,
            "changedCells": non_empty,
            "modelUsed": model_used,
            "durationSeconds": round(time.time() - started, 2),
        }
        add_history(entry)

        return jsonify({
            "ok": True,
            "rowNumber": row_number,
            "rowData": row_data,
            "changedCells": non_empty,
            "modelUsed": model_used,
            "historyEntry": entry,
        })
    except PermissionError as error:
        return api_error(
            RuntimeError("Close the Excel workbook first, then run extraction again."),
            status=423,
        )
    except Exception as error:
        return api_error(error)


@app.post("/api/open-sheet")
def api_open_sheet():
    try:
        payload = request.get_json(force=True) or {}
        open_file_on_host(str(payload.get("path", "")))
        return jsonify({"ok": True})
    except Exception as error:
        return api_error(error)


@app.get("/api/history")
def api_history():
    return jsonify({"ok": True, "history": read_history(50)})


@app.errorhandler(413)
def too_large(_error):
    return api_error(RuntimeError("File is too large. Use a file under 18 MB."), status=413)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="ChitraGuPT web app")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", default=5055, type=int)
    parser.add_argument("--open", action="store_true")
    return parser.parse_args()


def open_browser_after_start(host: str, port: int) -> None:
    browser_host = "127.0.0.1" if host in {"0.0.0.0", "::"} else host
    webbrowser.open(f"http://{browser_host}:{port}")


if __name__ == "__main__":
    args = parse_args()
    if args.open:
        threading.Timer(1.2, open_browser_after_start, args=(args.host, args.port)).start()

    print("ChitraGuPT")
    print(f"Local URL: http://127.0.0.1:{args.port}")
    if args.host == "0.0.0.0":
        print(f"Phone/LAN URL: http://{get_lan_ip()}:{args.port}")
    app.run(host=args.host, port=args.port, debug=False)
